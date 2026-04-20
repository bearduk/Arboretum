#!/usr/bin/env python3
"""
Export Keele arboretum Excel workbook to GeoJSON (EPSG:27700 → WGS84).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pyproj import Transformer

# British National Grid → WGS84 (lon, lat)
_TRANSFORMER = Transformer.from_crs("EPSG:27700", "EPSG:4326", always_xy=True)


def _warn(msg: str) -> None:
    print(msg, file=sys.stderr)


def header_to_key(cell) -> str | None:
    if cell is None or str(cell).strip() == "":
        return None
    s = str(cell).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or None


def _json_value(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        return v
    if isinstance(v, int):
        return v
    s = str(v).strip()
    return s if s else None


def _as_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except ValueError:
        return None


def _find_header_row(rows: list[tuple], marker: str = "tag") -> int:
    """0-based index of row whose first non-empty cell matches marker (as header start)."""
    for i, row in enumerate(rows):
        if not row:
            continue
        first = row[0]
        if first is not None and str(first).strip().lower() == marker:
            return i
    return 0


def sheet_to_features(
    rows: list[tuple],
    header_row_idx: int,
    source_sheet: str,
) -> tuple[list[dict], int, int]:
    """
    Returns (features, skipped_no_coords, skipped_empty).
    """
    header_row = rows[header_row_idx]
    keys = [header_to_key(h) for h in header_row]
    features: list[dict] = []
    skipped_no_coords = 0
    skipped_empty = 0
    tag_counts: Counter[str] = Counter()

    for row in rows[header_row_idx + 1 :]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            skipped_empty += 1
            continue

        props: dict = {"source_sheet": source_sheet}
        x_bng = None
        y_bng = None
        tag_val = None

        for j, key in enumerate(keys):
            if not key or j >= len(row):
                continue
            val = row[j]
            if key in ("x", "easting"):
                x_bng = _as_float(val)
                continue
            if key in ("y", "northing"):
                y_bng = _as_float(val)
                continue
            jv = _json_value(val)
            if jv is not None:
                props[key] = jv
            if key == "tag":
                tag_val = jv

        if tag_val is not None:
            tag_counts[str(tag_val)] += 1

        if x_bng is None or y_bng is None:
            skipped_no_coords += 1
            continue

        lon, lat = _TRANSFORMER.transform(x_bng, y_bng)
        props["easting"] = x_bng
        props["northing"] = y_bng

        features.append(
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                "properties": props,
            }
        )

    dupes = [t for t, c in tag_counts.items() if c > 1]
    if dupes:
        sample = ", ".join(sorted(dupes)[:15])
        more = f" (+{len(dupes) - 15} more)" if len(dupes) > 15 else ""
        _warn(
            f"[{source_sheet}] duplicate tag values ({len(dupes)} tags): {sample}{more}"
        )

    return features, skipped_no_coords, skipped_empty


def export_sheet(
    rows: list[tuple],
    header_row_idx: int | None,
    source_sheet: str,
    out_path: Path,
) -> None:
    if header_row_idx is None:
        header_row_idx = _find_header_row(rows, "tag")

    features, skip_nc, skip_em = sheet_to_features(rows, header_row_idx, source_sheet)
    fc = {"type": "FeatureCollection", "features": features}
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(fc, f, indent=2, ensure_ascii=False)
    _warn(
        f"Wrote {out_path.name}: {len(features)} features "
        f"(skipped {skip_nc} without X/Y, {skip_em} empty rows)"
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="XLSX → GeoJSON for Keele trees workbook.")
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        default=Path("Keele trees database.xlsx"),
        help="Path to Excel workbook",
    )
    parser.add_argument(
        "--out",
        "-o",
        type=Path,
        default=Path("docs/data"),
        help="Output directory for .geojson files",
    )
    args = parser.parse_args()

    if not args.input.is_file():
        _warn(f"Input not found: {args.input}")
        return 1

    wb = load_workbook(args.input, read_only=True, data_only=True)

    if "Cherries" in wb.sheetnames:
        ws = wb["Cherries"]
        rows = list(ws.iter_rows(values_only=True))
        export_sheet(rows, 0, "cherries", args.out / "cherries.geojson")

    if "other trees" in wb.sheetnames:
        ws = wb["other trees"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = _find_header_row(rows, "tag")
        export_sheet(rows, hdr, "other_trees", args.out / "other_trees.geojson")

    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
